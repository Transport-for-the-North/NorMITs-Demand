# -*- coding: utf-8 -*-
"""
    Module containing the test function for running the PopEmpComparator without
    running the whole EFS process.
"""

##### IMPORTS #####
# Standard imports
import os

# Third party imports
import pytest
import numpy as np
from openpyxl import Workbook

# Local imports
from pop_emp_comparator import PopEmpComparator, _excel_column_format
from demand_utilities import utils as du
from external_forecast_system import ExternalForecastSystem


##### CONSTANTS #####
# Constants for running the test function
MODEL_NAME = "norms_2015"
ITER_NUM = 1
IMPORT_LOC = "Y:/"
EXPORT_LOC = "Y:/"
BASE_YEAR = "2018"


##### FUNCTIONS #####
@pytest.mark.skip(reason="test on real data which can be run by running this module")
def test_real_data():
    """Tests the PopEmpComparator class on data from previous run of EFS. """
    imports, exports, _ = du.build_io_paths(
        IMPORT_LOC,
        EXPORT_LOC,
        MODEL_NAME,
        f"iter{ITER_NUM}",
        ExternalForecastSystem.__version__,
        ExternalForecastSystem._out_dir,
    )
    # Population csv files, locations from ExternalForecastSysten.__init__ parameters
    population_value_file = "population/base_population_2018.csv"
    population_growth_file = "population/future_population_growth.csv"
    population_constraint_file = "population/future_population_values.csv"
    # Employment csv files
    worker_value_file = "employment/base_workers_2018.csv"
    worker_growth_file = "employment/future_workers_growth.csv"
    worker_constraint_file = "employment/future_workers_growth_values.csv"

    # Compare the population inputs and outputs
    pop_comp = PopEmpComparator(
        os.path.join(imports["default_inputs"], population_value_file),
        os.path.join(imports["default_inputs"], population_growth_file),
        os.path.join(imports["default_inputs"], population_constraint_file),
        os.path.join(exports["productions"], "MSOA_population.csv"),
        "population",
        BASE_YEAR,
        msoa_lookup_file=os.path.join(imports["zoning"], "msoa_zones.csv"),
        sector_grouping_file=os.path.join(
            imports["zoning"], "tfn_sector_msoa_pop_weighted_lookup.csv"
        ),
    )
    pop_comp.write_comparisons(exports["reports"], output_as="csv", year_col=True)
    pop_comp.write_comparisons(exports["reports"], output_as="excel", year_col=True)
    # Compare the employment inputs and outputs
    emp_comp = PopEmpComparator(
        os.path.join(imports["default_inputs"], worker_value_file),
        os.path.join(imports["default_inputs"], worker_growth_file),
        os.path.join(imports["default_inputs"], worker_constraint_file),
        os.path.join(exports["attractions"], "MSOA_employment.csv"),
        "employment",
        BASE_YEAR,
        msoa_lookup_file=os.path.join(imports["zoning"], "msoa_zones.csv"),
        sector_grouping_file=os.path.join(
            imports["zoning"], "tfn_sector_msoa_emp_weighted_lookup.csv"
        ),
    )
    emp_comp.write_comparisons(exports["reports"], output_as="csv", year_col=True)
    emp_comp.write_comparisons(exports["reports"], output_as="excel", year_col=True)
    return


@pytest.mark.parametrize("style", ["Percent", "Comma [0]", None, "Normal"])
@pytest.mark.parametrize("ignore_rows", [0, 1])
def test_excel_column_format(style: str, ignore_rows: int):
    """Test the _excel_column_format function with different parameters.

    Parameters
    ----------
    style : str
        The style to convert column format to, passed to _excel_column_format.
    ignore_rows : int
        The number of rows to ignore, passed to _excel_column_format.
    """
    wb = Workbook()
    ws = wb.active
    rows = 2

    # Fill some cells
    for r in range(rows):
        ws.cell(row=r + 1, column=1, value=np.random.randint(100))

    _excel_column_format(ws, [style], ignore_rows=ignore_rows)

    style = "Normal" if style is None else style
    # Check styles
    for r in range(rows):
        this_style = "Normal" if r < ignore_rows else style
        new_style = ws.cell(row=r + 1, column=1).style
        assert new_style == this_style, f"'{new_style}' != '{this_style}' for row {r}"


##### MAIN #####
if __name__ == "__main__":
    test_real_data()
