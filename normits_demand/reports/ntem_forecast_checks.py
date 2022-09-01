# -*- coding: utf-8 -*-
"""
    Module containing functionality for providing summary spreadsheets
    for the NTEM forecast outputs.
"""

##### IMPORTS #####
# Standard imports
import functools
import re
from pathlib import Path
from typing import Dict, Any, Tuple, Union, List

# Third party imports
import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet import worksheet
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# Local imports
import normits_demand as nd
from normits_demand import core as nd_core
from normits_demand import logging as nd_log
from normits_demand.models.forecasting import ntem_forecast, tempro_trip_ends
from normits_demand.utils import file_ops, translation

##### CONSTANTS #####
LOG = nd_log.get_logger(__name__)


Matrix = Union[Path, pd.DataFrame]
"""Path to file (.csv or .pbz2), or DataFrame, containing matrix."""


##### FUNCTIONS #####
def _filename_contents(filename: str) -> Dict[str, Any]:
    """Extract information from matrix filenames.

    Parameters
    ----------
    filename : str
        Filename to extract information form,
        should not include file suffix.

    Returns
    -------
    Dict[str, Any]
        Dictionary containing matrix segmentation
        information with keys:
        - matrix_type
        - trip_end_type
        - year
        - purpose
        - nide

    Raises
    ------
    NTEMForecastError
        If `filename` isn't in the correct format.
    """
    pat = re.compile(
        "^"
        r"(?P<matrix_type>nhb|hb)"
        r"_(?P<trip_end_type>pa|od)"
        r"_yr(?P<year>\d{4})"
        r"_p(?P<purpose>\d{,2})"
        r"_m(?P<mode>\d)"
        "$",
        re.IGNORECASE,
    )
    match = pat.match(filename)
    if match is None:
        raise ntem_forecast.NTEMForecastError(
            f"filename ({filename!r}) is not in the correct format"
        )
    data = match.groupdict()
    for key in ("year", "purpose", "mode"):
        data[key] = int(data[key])
    return data


def _matrix_trip_ends(matrix: Matrix, trip_end_type: str) -> pd.DataFrame:
    """Calculate trip ends for a matrix file.

    Parameters
    ----------
    matrix : Matrix
        Path to matrix file or DataFrame.
    trip_end_type : str, {'pa', 'od'}
        Whether trip ends are productions and attractions
        or origins and destinations.

    Returns
    -------
    pd.DataFrame
        Trip end totals with 3 columns:
        - zone_id
        - trip_end_type
        - trips

    Raises
    ------
    NTEMForecastError
        If `trip_end_type` isn't 'pa' or 'od'.
    """
    trip_end_type = trip_end_type.lower().strip()
    if trip_end_type == "pa":
        te_names = ("attractions", "productions")
    elif trip_end_type == "od":
        te_names = ("destinations", "origins")
    else:
        raise ntem_forecast.NTEMForecastError(
            f"trip_end_type should be 'pa' or 'od' not {trip_end_type}"
        )
    if isinstance(matrix, (str, Path)):
        matrix = file_ops.read_df(matrix, index_col=0, find_similar=True)
    elif isinstance(matrix, pd.DataFrame):
        pass
    else:
        raise ntem_forecast.NTEMForecastError(f"matrix should be {Matrix} not {type(matrix)}")
    trip_ends = []
    for i, nm in enumerate(te_names):
        df = matrix.sum(axis=i)
        df.index.name = "zone_id"
        df = df.to_frame(name="trips")
        df.insert(0, "trip_end_type", nm)
        trip_ends.append(df.reset_index())
    trip_ends = pd.concat(trip_ends, axis=0)
    trip_ends.loc[:, "zone_id"] = pd.to_numeric(trip_ends["zone_id"], downcast="integer")
    return trip_ends


def _compare_trip_ends(
    base_trip_ends: Dict[str, Dict[str, nd_core.DVector]],
    forecast_trip_ends: Dict[str, Dict[str, nd_core.DVector]],
    tempro_data: tempro_trip_ends.TEMProTripEnds,
    years: Tuple[int, int],
    comparison_zone_system: dict[str, str],
) -> pd.DataFrame:
    """Compares matrix growth to `tempro_data`.

    Internal functionality for `pa_matrix_comparison`.
    """
    tempro_comparison = []
    for mat_type in base_trip_ends:
        if mat_type not in forecast_trip_ends:
            raise KeyError(f"{mat_type!r} key not in forecast_trip_ends")
        for te_type in base_trip_ends[mat_type]:
            if te_type not in forecast_trip_ends[mat_type]:
                raise KeyError(f"{te_type!r} not in forecast_trip_ends[{mat_type!r}]")
            dvectors = {
                f"matrix_{years[0]}": base_trip_ends[mat_type][te_type],
                f"matrix_{years[1]}": forecast_trip_ends[mat_type][te_type],
            }
            # Check TEMPro DVector
            for y in years:
                dvec = getattr(tempro_data, f"{mat_type}_{te_type}")[y]
                base_dvec = dvectors[f"matrix_{years[0]}"]
                if dvec.segmentation != base_dvec.segmentation:
                    raise ntem_forecast.NTEMForecastError(
                        "TEMPro trip ends segmentation should be "
                        f"{base_dvec.segmentation.name} not "
                        f"{dvec.segmentation.name}"
                    )
                if dvec.zoning_system != base_dvec.zoning_system:
                    raise ntem_forecast.NTEMForecastError(
                        "TEMPro trip ends zoning system should be "
                        f"{base_dvec.zoning_system.name} not "
                        f"{dvec.zoning_system.name}"
                    )
                dvectors[f"tempro_{y}"] = dvec

            # Convert to DataFrames
            dataframes = []
            index_cols = ["p", "m", f"{comparison_zone_system['trip end']}_zone_id"]
            for nm, dvec in dvectors.items():
                df = dvec.to_df().rename(columns={"val": nm})
                df = df.set_index(index_cols)
                dataframes.append(df)
            df = pd.concat(dataframes, axis=1)
            df = df.reset_index()
            df.loc[:, "matrix_type"] = mat_type
            df.loc[:, "trip_end_type"] = te_type
            tempro_comparison.append(
                df.set_index(["matrix_type", "trip_end_type"] + index_cols)
            )
    tempro_comparison = pd.concat(tempro_comparison)
    # Calculate growth differences
    tempro_comparison.loc[:, "matrix_growth"] = (
        tempro_comparison[f"matrix_{years[1]}"] / tempro_comparison[f"matrix_{years[0]}"]
    )
    tempro_comparison.loc[:, "tempro_growth"] = (
        tempro_comparison[f"tempro_{years[1]}"] / tempro_comparison[f"tempro_{years[0]}"]
    )
    tempro_comparison.loc[:, "growth_difference"] = (
        tempro_comparison["matrix_growth"] - tempro_comparison["tempro_growth"]
    )
    tempro_comparison.loc[:, "growth_%_diff"] = (
        tempro_comparison["matrix_growth"] / tempro_comparison["tempro_growth"]
    ) - 1
    return tempro_comparison


def matrix_dvectors(
    matrices: Dict[int, Matrix],
    segmentation: str,
    trip_end_type: str,
    matrix_zoning: str,
    mode: int,
    comparison_zone_systems: dict[str, str],
) -> Dict[str, nd_core.DVector]:
    """Calculate matrix trip ends and convert to DVectors

    Parameters
    ----------
    matrices : Dict[int, Matrix]
        Paths to, or DataFrames of, the matrices by
        purpose (keys), the files should be either
        '.csv' or '.pbz2'.
    segmentation : str
        Name of the segmentation level for the
        returned DVectors.
    trip_end_type : str, {'pa', 'od'}
        The type of matrices being used.
    matrix_zoning : str
        Name of the zone system of the matrix.
    mode : int
        The mode number of the matrix.

    Returns
    -------
    Dict[str, nd_core.DVector]
        DVectors for both trip ends (productions and attractions
        or origins and destinations). DVectors have the segmentation
        given and the zone system defined by `COMPARISON_ZONE_SYSTEM`.
    """
    # Get matrix trip ends
    trip_ends = []
    for p, mat in matrices.items():
        df = _matrix_trip_ends(mat, trip_end_type)
        df.loc[:, "p"] = p
        df.loc[:, "m"] = mode
        trip_ends.append(df)
    trip_ends = pd.concat(trip_ends)

    matrix_zoning = nd_core.get_zoning_system(matrix_zoning)
    comparison_zoning = nd_core.get_zoning_system(comparison_zone_systems["trip end"])
    dvectors = {}
    columns = ["zone_id", "trips", "p", "m"]
    for te_type in trip_ends.trip_end_type.unique():
        # Convert trip ends to DVector
        dvec = nd_core.DVector(
            nd_core.get_segmentation_level(segmentation),
            trip_ends.loc[trip_ends.trip_end_type == te_type, columns],
            matrix_zoning,
            time_format="avg_day",
            zone_col="zone_id",
            val_col="trips",
            df_naming_conversion={"p": "p", "m": "m"},
        )
        dvectors[te_type] = dvec.translate_zoning(comparison_zoning)
    return dvectors


def _find_matrices(folder: Path) -> pd.DataFrame:
    """Find matrix files in given `folder`.

    Finds all files with extension ('.pbz2' or '.csv')
    and checks the filename is the correct format.

    Parameters
    ----------
    folder : Path
        Path to matrix folder.

    Returns
    -------
    pd.DataFrame
        Paths to all the matrices found with information
        extracted from the filename, contains column 'path'
        with the file paths and index columns:
        - `matrix_type`: either 'hb' or 'nhb'
        - `year`
        - `mode`
        - `purpose`
    """
    files = []
    file_types = (".pbz2", ".csv", ".csv.bz2")
    for p in folder.iterdir():
        suffixes = "".join(p.suffixes)
        if p.is_dir() or suffixes.lower() not in file_types:
            continue
        suffix = "".join(p.suffixes)
        try:
            file_data = _filename_contents(p.name.removesuffix(suffix))
        except ntem_forecast.NTEMForecastError as err:
            LOG.warning(err)
            continue
        file_data["path"] = p
        files.append(file_data)
    files_df = pd.DataFrame(files)
    files_df.to_csv(folder / "Matrices list.csv", index=False)
    index_cols = ["matrix_type", "year", "mode", "purpose"]
    files_df = files_df.loc[:, index_cols + ["path"]].set_index(index_cols)
    return files_df


def _read_matrices(paths: Dict[int, Path]) -> Dict[int, pd.DataFrame]:
    """Reads matrices and returns dictionary of DataFrames with the same keys."""
    matrices = {}
    for i, p in paths.items():
        df = file_ops.read_df(p, index_col=0, find_similar=True)
        df.columns = pd.to_numeric(df.columns, downcast="integer", errors="ignore")
        matrices[i] = df
    return matrices


def pa_matrix_comparison(
    ntem_imports: ntem_forecast.NTEMImportMatrices,
    pa_folder: Path,
    tempro_data: tempro_trip_ends.TEMProTripEnds,
    mode: nd.Mode,
    comparison_zone_system: dict[str, str],
    base_year: int,
):
    """Produce TEMPro comparisons for PA matrices.

    Parameters
    ----------
    ntem_imports : ntem_forecast.NTEMImportMatrices
        Paths to the input Post ME base matrices.
    pa_folder : Path
        Folder containing PA matrices.
    tempro_data : TEMProTripEnds
        TEMPro trip end data.
    base_year : int
        Base model year.
    """
    LOG.info("PA matrix trip ends comparison with TEMPro")
    output_folder = pa_folder / "TEMPro Comparisons"
    output_folder.mkdir(exist_ok=True)
    # Extract information from filenames
    files = _find_matrices(pa_folder)

    # Read base matrices
    if ntem_imports.mode != 3:
        raise NotImplementedError("PA matrix comparison only works for mode 3")
    segmentation = {"hb": f"hb_p_m_{mode.name.lower()}", "nhb": f"nhb_p_m_{mode.name.lower()}"}
    base_matrices = {}
    base_trip_ends = {}
    for nm, seg in segmentation.items():
        LOG.info("Getting trip ends for base %s", nm.upper())
        base_matrices[nm] = _read_matrices(getattr(ntem_imports, f"{nm}_paths"))
        base_trip_ends[nm] = matrix_dvectors(
            base_matrices[nm],
            seg,
            "pa",
            ntem_imports.model.get_zoning_system().name,
            ntem_imports.mode,
            comparison_zone_system,
        )

    # Convert tempro_data to LA zoning and make sure segmentation is (n)hb_p_m
    tempro_data_comp = tempro_data.translate_zoning(comparison_zone_system["trip end"])
    # Compare trip ends to tempro for all purposes and years
    for yr in files.index.get_level_values("year").unique():
        forecast_matrices = {}
        forecast_trip_ends = {}
        for nm, seg in segmentation.items():
            LOG.info("Getting trip ends for %s %s", yr, nm.upper())
            indices = pd.IndexSlice[nm, yr, ntem_imports.mode]
            forecast_matrices[nm] = _read_matrices(files.loc[indices, "path"].to_dict())
            forecast_trip_ends[nm] = matrix_dvectors(
                forecast_matrices[nm],
                seg,
                "pa",
                ntem_imports.model.get_zoning_system().name,
                ntem_imports.mode,
                comparison_zone_system,
            )
        comparison = _compare_trip_ends(
            base_trip_ends,
            forecast_trip_ends,
            tempro_data_comp,
            (base_year, yr),
            comparison_zone_system=comparison_zone_system,
        )
        out = (
            output_folder
            / f"PA_TEMPro_comparisons-{yr}-{comparison_zone_system['trip end']}.csv"
        )
        file_ops.write_df(comparison, out)
        LOG.info("Written: %s", out)

        for nm, comp_zone in comparison_zone_system.items():
            if not nm.startswith("matrix"):
                continue
            LOG.info("Matrix comparisons at %s zoning", comp_zone)
            matrix_comparison(
                base_matrices,
                forecast_matrices,
                ntem_imports.model.get_zoning_system().name,
                tempro_data,
                comp_zone,
                (base_year, yr),
                output_folder / f"PA_TEMPro_comparisons-{yr}-{comp_zone}",
            )


def translate_matrix(
    matrix: pd.DataFrame,
    matrix_zoning_name: str,
    new_zoning_name: str,
    weighting: str = None,
    **kwargs,
) -> pd.DataFrame:
    """Tranlate square matrix into new zoning system.

    Wrapper for `translation.pandas_matrix_zone_translation`.

    Parameters
    ----------
    matrix : pd.DataFrame
        Matrix in square format i.e. column names
        and row indices are the zones.
    matrix_zoning_name : str
        Name of the current zone system.
    new_zoning_name : str
        Name of the zone system to translate to.
    weighting : str, optional
        Translation weighting to use.

    Returns
    -------
    pd.DataFrame
        `matrix` after translation to `new_zoning_name`.
    """
    # BACKLOG: Move this function to a better location
    #   labels: QoL Updates
    if matrix_zoning_name == new_zoning_name:
        return matrix
    # Get correspondence DataFrame
    matrix_zoning = nd_core.get_zoning_system(matrix_zoning_name)
    new_zoning = nd_core.get_zoning_system(new_zoning_name)
    lookup = matrix_zoning._get_translation_definition(new_zoning, weighting)
    # Translate matrix
    return translation.pandas_matrix_zone_translation(
        matrix,
        f"{matrix_zoning_name}_zone_id",
        f"{new_zoning_name}_zone_id",
        f"{matrix_zoning_name}_to_{new_zoning_name}",
        matrix_zoning.unique_zones,
        new_zoning.unique_zones,
        translation=lookup,
        **kwargs,
    )


def _matrix_comparison_write(
    tempro_data: pd.DataFrame,
    base_matrices: pd.DataFrame,
    forecast_matrices: pd.DataFrame,
    output_path: Path,
    years: Tuple[int, int],
    zone_col: str,
):
    """Write the `matrix_comparison` outputs to an Excel file."""
    TEMPRO_COLUMNS = [
        "matrix_type",
        "trip_end_type",
        "p",
        "m",
        zone_col,
        "id",
        *[f"tempro_{y}" for y in years],
    ]
    MATRIX_COLUMNS = [
        "matrix_type",
        "purpose",
        "from_zone",
        "to_zone",
        "id",
        "trips",
    ]
    MATRIX_SHEETS = [f"{s} Matrices Data" for s in ("Base", "Forecast")]
    out = output_path.with_suffix(".xlsx")
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        tempro_data[TEMPRO_COLUMNS].to_excel(writer, sheet_name="TEMPro Data", index=False)
        for nm, mat in zip(MATRIX_SHEETS, (base_matrices, forecast_matrices)):
            mat[MATRIX_COLUMNS].to_excel(writer, sheet_name=nm, index=False)

        # Create summary sheet
        wb: openpyxl.Workbook = writer.book
        ws = wb.create_sheet("Summary", 0)
        # Add purpose dropdown
        purposes = base_matrices["purpose"].unique().tolist()
        valid_purp = DataValidation(
            "list",
            formula1=f'"{",".join(str(p) for p in purposes)}"',
        )
        ws.add_data_validation(valid_purp)
        ws["B2"] = "Purpose"
        PURP_CELL = "C2"
        ws[PURP_CELL] = purposes[0]
        valid_purp.add(PURP_CELL)

        zones = tempro_data[zone_col].unique().tolist()
        endrow = 3
        for nm in MATRIX_SHEETS:
            endrow, _ = _excel_matrix_formula(ws, nm, endrow + 2, 2, zones, PURP_CELL)
        _excel_growth_matrix(ws, endrow + 2, 2, zones)

    LOG.info("Written: %s", out)


def _excel_matrix_formula(
    ws: worksheet.Worksheet,
    name: str,
    startrow: int,
    startcol: int,
    zones: List[Any],
    purpose_cell: str,
) -> Tuple[int, int]:
    """Add summary matrix tables for `_matrix_comparison_write`."""
    index_formula = r"""
        =INDEX('{matrix_sheet}'!F:F,
            MATCH(
                {purp_cell}&"_"&{row_cell}&"_"&{col_cell},
                '{matrix_sheet}'!E:E,
                0
            )
        )
    """
    total_formula = "=SUM({start}:{end})"
    tempro_formula = r"""
        =INDEX('{tempro_sheet}'!{data}:{data},
            MATCH(
                "{pa}_"&{purp_cell}&"_"&{row_cell},
                '{tempro_sheet}'!F:F,
                0
            )
        )
    """
    # Replace whitespace for ease of reading within Excel
    index_formula = re.sub(r"\s+", "", index_formula)
    tempro_formula = re.sub(r"\s+", "", tempro_formula)
    TEMPRO_SHEET = "TEMPro Data"
    if name.lower().startswith("base"):
        tempro_data_column = "G"
    else:
        tempro_data_column = "H"

    cell_id = lambda r, c: f"{get_column_letter(c)}{r}"
    ws.cell(startrow, startcol, name)
    for row, rzone in enumerate(zones, 1):
        ws.cell(startrow + row, startcol, rzone)
        for col, czone in enumerate(zones, 1):
            if row == 1:
                ws.cell(startrow, startcol + col, czone)
            ws.cell(
                startrow + row,
                startcol + col,
                index_formula.format(
                    matrix_sheet=name,
                    purp_cell=purpose_cell,
                    row_cell=cell_id(startrow + row, startcol),
                    col_cell=cell_id(startrow, startcol + col),
                ),
            )
        # Add total column and TEMPro data column
        col = len(zones) + 1
        if row == 1:
            ws.cell(startrow, startcol + col, "Total")
            ws.cell(startrow, startcol + col + 1, "TEMPro Productions")
        ws.cell(
            startrow + row,
            startcol + col,
            total_formula.format(
                start=cell_id(startrow + row, startcol + 1),
                end=cell_id(startrow + row, startcol + col - 1),
            ),
        )
        ws.cell(
            startrow + row,
            startcol + col + 1,
            tempro_formula.format(
                tempro_sheet=TEMPRO_SHEET,
                data=tempro_data_column,
                pa="productions",
                purp_cell=purpose_cell,
                row_cell=cell_id(startrow + row, startcol),
            ),
        )

    # Add total row and TEMPro data attractions
    row = len(zones) + 1
    ws.cell(startrow + row, startcol, "Total")
    ws.cell(startrow + row + 1, startcol, "TEMPro Attractions")
    for col in range(1, len(zones) + 3):
        ws.cell(
            startrow + row,
            startcol + col,
            total_formula.format(
                start=cell_id(startrow + 1, startcol + col),
                end=cell_id(startrow + row - 1, startcol + col),
            ),
        )
        if col <= len(zones):
            ws.cell(
                startrow + row + 1,
                startcol + col,
                tempro_formula.format(
                    tempro_sheet=TEMPRO_SHEET,
                    data=tempro_data_column,
                    pa="attractions",
                    purp_cell=purpose_cell,
                    row_cell=cell_id(startrow, startcol + col),
                ),
            )
        elif col == len(zones) + 1:
            # Calculate TEMPro row total
            ws.cell(
                startrow + row + 1,
                startcol + col,
                total_formula.format(
                    start=cell_id(startrow + row + 1, startcol + 1),
                    end=cell_id(startrow + row + 1, startcol + col - 1),
                ),
            )
        elif col == len(zones) + 2:
            # Calculate TEMPro total
            ws.cell(
                startrow + row + 1,
                startcol + col,
                "={c1}+{c2}".format(
                    c1=cell_id(startrow + row + 1, startcol + col - 1),
                    c2=cell_id(startrow + row, startcol + col),
                ),
            )
    return startrow + row + 1, startcol + col + 1


def _excel_growth_matrix(
    ws: worksheet.Worksheet, startrow: int, startcol: int, zones: List[Any]
):
    """Add growth matrix summary for `_matrix_comparison_write`."""
    cell_id = lambda r, c: f"{get_column_letter(c)}{r}"
    ws.cell(startrow, startcol, "Growth")
    row_list = zones + ["Total", "TEMPro"]
    for row, rzone in enumerate(row_list, 1):
        ws.cell(startrow + row, startcol, rzone)
        for col, czone in enumerate(row_list, 1):
            if row == 1:
                ws.cell(startrow, startcol + col, czone)
            row_diff = len(row_list) + 2
            base_pos = cell_id(startrow + row - (2 * row_diff), startcol + col)
            forecast_pos = cell_id(startrow + row - row_diff, startcol + col)
            ws.cell(startrow + row, startcol + col, f"={forecast_pos}/{base_pos}")


def matrix_comparison(
    base_matrices: Dict[str, Dict[int, pd.DataFrame]],
    forecast_matrices: Dict[str, Dict[int, pd.DataFrame]],
    matrix_zoning: str,
    tempro_data: tempro_trip_ends.TEMProTripEnds,
    comparison_zoning: str,
    years: Tuple[int, int],
    output_path: Path,
):
    """Convert matrices to summary sector systems for comparisons.

    Produces Excel files containing comparisons to TEMPro at
    different sector systems.

    Parameters
    ----------
    base_matrices : Dict[str, Dict[int, pd.DataFrame]]
        Base matrices for all purposes where keys for
        the first dictionary are 'hb' or 'nhb' and the
        second dictionary is the purpose number.
    forecast_matrices : Dict[str, Dict[int, pd.DataFrame]]
        Forecast matrices for all purposes where keys for
        the first dictionary are 'hb' or 'nhb' and the
        second dictionary is the purpose number.
    matrix_zoning : str
        Name of the matrix zone system.
    tempro_data : tempro_trip_ends.TEMProTripEnds
        TEMPro trip end data for comparisons.
    comparison_zoning : str
        Name of the zone system to convert to
        for comparison.
    years : Tuple[int, int]
        The base and forecast years.
    output_path : Path
        Excel file to save output to.
    """
    # Translate matrices and tempro data to comparison_zoning
    tempro_data = tempro_data.translate_zoning(comparison_zoning)
    zone_col = f"{comparison_zoning}_zone_id"
    mat_translation = functools.partial(
        translate_matrix, matrix_zoning_name=matrix_zoning, new_zoning_name=comparison_zoning
    )
    # Translate matrices and convert to long format with
    # columns for matrix type and purpose
    long_matrices = {"base": [], "forecast": []}
    matrix_iterator = tuple(zip(long_matrices, (base_matrices, forecast_matrices)))
    tempro_df = []
    for mat_type in base_matrices:
        # Extract both trip ends from tempro for all purposes
        for pa in ("productions", "attractions"):
            temp_tempro = []
            for yr in years:
                df = getattr(tempro_data, f"{mat_type}_{pa}")[yr]
                df = df.to_df().rename(columns={"val": f"tempro_{yr}"})
                df.loc[:, "matrix_type"] = mat_type
                df.loc[:, "trip_end_type"] = pa
                temp_tempro.append(
                    df.set_index(["matrix_type", "trip_end_type", "p", "m", zone_col])
                )
            tempro_df.append(pd.concat(temp_tempro, axis=1).reset_index())
        # Convert each matrix to long format and add information
        # about mat type and purpose
        for p in base_matrices[mat_type]:
            for nm, mat in matrix_iterator:
                df = mat_translation(mat[mat_type][p]).stack().to_frame("trips")
                df.index.names = ["from_zone", "to_zone"]
                df.loc[:, "matrix_type"] = mat_type
                df.loc[:, "purpose"] = p
                long_matrices[nm].append(df.reset_index())

    # Concatenate TEMPro data then groupby the index to avoid duplicate
    # indices with NaNs in the empty year columns
    tempro_df = pd.concat(tempro_df)
    tempro_df.loc[:, "id"] = (
        tempro_df["trip_end_type"].astype(str)
        + "_"
        + tempro_df["p"].astype(str)
        + "_"
        + tempro_df[zone_col].astype(str)
    )
    for nm, ls in long_matrices.items():
        df = pd.concat(ls)
        df.loc[:, "id"] = (
            df["purpose"].astype(str)
            + "_"
            + df["from_zone"].astype(str)
            + "_"
            + df["to_zone"].astype(str)
        )
        long_matrices[nm] = df[
            ["matrix_type", "purpose", "from_zone", "to_zone", "id", "trips"]
        ]

    # Save data to Excel
    _matrix_comparison_write(
        tempro_df,
        long_matrices["base"],
        long_matrices["forecast"],
        output_path,
        years,
        zone_col,
    )


def od_matrix_comparison(
    base_folder: Path,
    forecast_folder: Path,
    matrix_zoning: str,
    comparison_zoning: str,
    user_classes: list[str],
    time_periods: list[int],
    future_years: list[int],
):
    """Write spreadsheet summarising OD matrix growth.

    Parameters
    ----------
    base_folder : Path
        Folder containing base OD matrices.
    forecast_folder : Path
        Folder containing forecast OD matrices.
    matrix_zoning : str
        Name of the current matrix zoning system.
    comparison_zoning : str
        Name of the zoning system for the summaries.
    years : List[int]
        List of forecast years.
    """
    # TODO(MB) Make this function more robust for base matrix names

    OD_MATRIX_NAMES = {
        "base": "od_m3_{purp}_tp{tp}_postME.csv",
        "base2": "od_m3_{purp}_tp{tp}.csv",
        "forecast": "od_{purp}_yr{yr}_m3_tp{tp}.csv",
    }
    for name, folder in (("base", base_folder), ("forecast", forecast_folder)):
        if not folder.is_dir():
            raise NotADirectoryError(f"{name} folder doesn't exist: {folder}")
    LOG.info(
        "Creating OD matrix summary using:\n\t"
        "base folder: %s\n\tforecast folder: %s\n\t"
        "at zoning system: %s",
        base_folder,
        forecast_folder,
        comparison_zoning,
    )
    out_path = forecast_folder / f"OD_matrix_growth_summary-{comparison_zoning}.xlsx"
    pbar = tqdm(
        desc="Creating OD Matrix Summary",
        total=len(user_classes) * len(time_periods),
        dynamic_ncols=True,
    )
    with pd.ExcelWriter(out_path) as writer:
        for purpose in user_classes:
            for tp in time_periods:
                base_path = base_folder / OD_MATRIX_NAMES["base"].format(purp=purpose, tp=tp)
                if not base_path.exists():
                    print(base_path)
                    base_path = base_path.with_name(
                        OD_MATRIX_NAMES["base2"].format(purp=purpose, tp=tp)
                    )
                forecast_name = lambda p, t, y: OD_MATRIX_NAMES["forecast"].format(
                    purp=p, tp=t, yr=y
                )
                forecast_paths = {
                    y: forecast_folder / forecast_name(purpose, tp, y) for y in future_years
                }
                _compare_od_matrices(
                    writer,
                    f"{purpose} TP{tp}",
                    base_path,
                    forecast_paths,
                    matrix_zoning,
                    comparison_zoning,
                )
                pbar.update()
    pbar.close()
    LOG.info("Written: %s", out_path)


def _compare_od_matrices(
    excel_writer,
    sheet: str,
    base_path: Path,
    forecast_paths: Dict[int, Path],
    matrix_zoning: str,
    comparison_zoning: str,
):
    """Writes a single summary sheet to the spreadsheet.

    Internal function for `od_matrix_comparison`.
    """

    def matrix_totals(matrix: pd.DataFrame) -> pd.DataFrame:
        """Add total row and column to matrix."""
        matrix.loc[:, "Total"] = matrix.sum(axis=1)
        matrix.loc["Total"] = matrix.sum(axis=0)
        return matrix

    # Read base matrix which is in long format
    base = file_ops.read_df(base_path, index_col=0)
    if base.isnull().values.any():
        LOG.warning(
            "Base matrix at %s contains %s null values.  These are being"
            " set to zero internally for reporting but consider checking the matrix.",
            base_path,
            base.isnull().sum().sum(),
        )
        base.fillna(0, inplace=True)
    base.rename(columns={i: int(i) for i in base.columns}, inplace=True)
    base = translate_matrix(base, matrix_zoning, comparison_zoning)
    base = matrix_totals(base)
    base.to_excel(excel_writer, sheet_name=sheet, index_label="Base")

    for i, (yr, path) in enumerate(forecast_paths.items()):
        # Read forecast matrix which is in square format
        forecast = file_ops.read_df(path, index_col=0)
        forecast.columns = pd.to_numeric(forecast.columns, downcast="integer")
        if forecast.isnull().values.any():
            LOG.warning(
                "Forecast matrix at %s contains %s null values.  These are "
                "set to zero internally for reporting but consider checking the matrix.",
                path,
                forecast.isnull().sum().sum(),
            )
            forecast.fillna(0, inplace=True)
        forecast = translate_matrix(forecast, matrix_zoning, comparison_zoning)
        forecast = matrix_totals(forecast)
        col = i * (len(forecast) + 2)
        forecast.to_excel(
            excel_writer,
            sheet_name=sheet,
            index_label=f"Forecast - {yr}",
            startrow=len(base) + 3,
            startcol=col,
        )
        growth = forecast / base
        growth.to_excel(
            excel_writer,
            sheet_name=sheet,
            index_label=f"Growth - {yr}",
            startrow=2 * (len(base) + 3),
            startcol=col,
        )
