# -*- coding: utf-8 -*-
"""
Created on: Mon June 08:12:21 2021
Updated on:

Original author: Ben Taylor
Last update made by:
Other updates made by:

File purpose:
Collection of utility functions specifically for manipulating pandas
"""
# Builtins
import os
import operator

from typing import Any
from typing import List
from typing import Callable
from typing import Iterable

# Third Party
import numpy as np
import pandas as pd
import openpyxl

from openpyxl.utils import dataframe as openpyxl_dataframe

# Local

def get_wide_mask(df: pd.DataFrame,
                  zones: List[Any] = None,
                  col_zones: List[Any] = None,
                  index_zones: List[Any] = None,
                  join_fn: Callable = operator.and_
                  ) -> np.ndarray:
    """
    Generates a mask for a wide matrix. Returned mask will be same shape as df

    The zones the set the mask for can be set individually with col_zones and
    index_zones, or to the same value with zones.


    Parameters
    ----------
    df:
        The dataframe to generate the mask for

    zones:
        The zones to match to in both the columns and index. If this value
        is set it will overwrite anything passed into col_zones and
        index_zones.

    col_zones:
        The zones to match to in the columns. This value is ignored if
        zones is set.

    index_zones:
        The zones to match to in the index. This value is ignored if
        zones is set.

    join_fn:
        The function to call on the column and index masks to join them.
        By default, a bitwise and is used. See pythons builtin operator
        library for more options.

    Returns
    -------
    mask:
        A mask of true and false values. Will be the same shape as df.
    """
    # Validate input args
    if zones is None:
        if col_zones is None or index_zones is None:
            raise ValueError(
                "If zones is not set, both col_zones and row_zones need "
                "to be set."
            )
    else:
        col_zones = zones
        index_zones = zones

    # Try and cast to the correct types for rows/cols
    try:
        # Assume columns are strings if they are an object
        col_dtype = df.columns.dtype
        col_dtype = str if col_dtype == object else col_dtype
        col_zones = np.array(col_zones, col_dtype)
    except ValueError:
        raise ValueError(
            "Cannot cast the col_zones to the required dtype to match the "
            "dtype of the given df columns. Tried to cast to: %s"
            % str(df.columns.dtype)
        )

    try:
        index_zones = np.array(index_zones, df.index.dtype)
    except ValueError:
        raise ValueError(
            "Cannot cast the index_zones to the required dtype to match the "
            "dtype of the given df index. Tried to cast to: %s"
            % str(df.index.dtype)
        )

    # Create square masks for the rows and cols
    col_mask = np.broadcast_to(df.columns.isin(col_zones), df.shape)
    index_mask = np.broadcast_to(df.index.isin(index_zones), df.shape).T

    # Combine to get the full mask
    return join_fn(col_mask, index_mask)


def get_internal_mask(df: pd.DataFrame,
                      zones: List[Any],
                      ) -> np.ndarray:
    """
    Generates a mask for a wide matrix. Returned mask will be same shape as df

    Parameters
    ----------
    df:
        The dataframe to generate the mask for

    zones:
        A list of zone numbers that make up the internal zones

    Returns
    -------
    mask:
        A mask of true and false values. Will be the same shape as df.
    """
    return get_wide_mask(df=df, zones=zones, join_fn=operator.and_)


def get_external_mask(df: pd.DataFrame,
                      zones: List[Any],
                      ) -> np.ndarray:
    """
    Generates a mask for a wide matrix. Returned mask will be same shape as df

    Parameters
    ----------
    df:
        The dataframe to generate the mask for

    zones:
        A list of zone numbers that make up the external zones

    Returns
    -------
    mask:
        A mask of true and false values. Will be the same shape as df.
    """
    return get_wide_mask(df=df, zones=zones, join_fn=operator.or_)


def get_external_values(df: pd.DataFrame,
                        zones: List[Any],
                        ) -> pd.DataFrame:
    """Get only the external values in df

    External values contains internal-external, external-internal, and
    external-external. All values not meeting this criteria will be set
    to 0.

    Parameters
    ----------
    df:
        The dataframe to get the external values from

    zones:
        A list of zone numbers that make up the external zones

    Returns
    -------
    external_df:
        A dataframe containing only the external demand from df.
        Will be the same shape as df.
    """
    return df * get_external_mask(df, zones)


def get_internal_values(df: pd.DataFrame,
                        zones: List[Any],
                        ) -> pd.DataFrame:
    """Get only the internal values in df

    Internal values contains internal-internal. All values not
    meeting this criteria will be set to 0.

    Parameters
    ----------
    df:
        The dataframe to get the external values from

    zones:
        A list of zone numbers that make up the internal zones

    Returns
    -------
    internal_df:
        A dataframe containing only the internal demand from df.
        Will be the same shape as df.
    """
    return df * get_internal_mask(df, zones)


def internal_external_report(df: pd.DataFrame,
                             internal_zones: List[Any],
                             external_zones: List[Any],
                             ) -> pd.DataFrame:
    """Generates a report df of values in internal/external zones

    Generates a dataframe with 4 rows, each showing the total across
    that portion of the matrix. The dataframe is split into:
    internal-internal
    internal-external
    external-internal
    external-external

    Parameters
    ----------
    df:
        The dataframe to generate the report on.

    internal_zones:
        A list of the internal zones of the zoning system used by df

    external_zones
        A list of the external zones of the zoning system used by df

    Returns
    -------
    report:
        A report of internal and external demand in df.
    """
    # Build the initial report
    index = pd.Index(['internal', 'external'])
    report = pd.DataFrame(
        index=index,
        columns=index,
        data=np.zeros((len(index), len(index)))
    )

    # Build the kwargs to iterate over
    report_kwargs = {
        ('internal', 'internal'): {'index_zones': internal_zones, 'col_zones': internal_zones},
        ('internal', 'external'): {'index_zones': internal_zones, 'col_zones': external_zones},
        ('external', 'internal'): {'index_zones': external_zones, 'col_zones': internal_zones},
        ('external', 'external'): {'index_zones': external_zones, 'col_zones': external_zones},
    }

    # Build the report from the kwargs
    for (row_idx, col_idx), kwargs in report_kwargs.items():
        # Pull out just the trips for this section
        mask = get_wide_mask(
            df=df,
            join_fn=operator.and_,
            **kwargs,
        )
        total = (df * mask).values.sum()

        # Feel like this indexing is backwards...
        report[col_idx][row_idx] = total

    # Add a total row and column
    report['total'] = report.values.sum(axis=1)
    report.loc['total'] = report.values.sum(axis=0)

    return report


def _openpyxl_df_to_excel(
    df: pd.DataFrame,
    path: os.PathLike,
    sheet_name: str = "Sheet1",
    header: bool = True,
    index: bool = True,
    start_row: int = 0,
    start_col: int = 0,
) -> None:
    """Append a DataFrame to existing Excel workbook

    Append `df` to existing Excel file `path` into `sheet_name` Sheet.
    `start_row` and `start_col` are used to define the cell that the top left
    corner of `df` will begin.
    If `path` doesn't exist, then it is created.

    Parameters
    ----------
    df:
        DataFrame to save to workbook

    path:
        Path to the Excel workbook to write to.

    sheet_name:
        Name of the sheet to save `df`.

    header:
        Whether to write out the column names.

    index:
        Whether to write out the index (row) names.

    start_row:
        Upper left cell row to dump `df`.

    start_col:
        Upper left cell col to dump `df`.

    Returns
    -------
    None
    """
    # Create the workbook if it doesn't exist
    if os.path.isfile(path):
        workbook = openpyxl.load_workbook(path)
    else:
        workbook = openpyxl.Workbook()

    # Grab the sheet to write to
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Write each cell of df into individual cells
    df_rows = openpyxl_dataframe.dataframe_to_rows(df, index=index, header=header)
    for r_idx, row in enumerate(df_rows, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(
                row=start_row + r_idx,
                column=start_col + c_idx,
                value=value,
            )

    workbook.save(filename=path)


def append_df_to_excel(
    df: pd.DataFrame,
    path: os.PathLike,
    sheet_name: str = "Sheet1",
    header: bool = True,
    index: bool = True,
    start_row: int = 0,
    start_col: int = 0,
    keep_data_validation: bool = False,
    **to_excel_kwargs,
) -> None:
    """Append a DataFrame to existing Excel workbook

    Append `df` to existing Excel file `path` into `sheet_name` Sheet.
    `start_row` and `start_col` are used to define the cell that the top left
    corner of `df` will begin.
    If `path` doesn't exist, then it is created.

    Parameters
    ----------
    df:
        DataFrame to save to workbook

    path:
        Path to the Excel workbook to write to.

    sheet_name:
        Name of the sheet to save `df`.

    header:
        Whether to write out the column names.

    index:
        Whether to write out the index (row) names.

    start_row:
        Upper left cell row to dump `df`.

    start_col:
        Upper left cell col to dump `df`.

    keep_data_validation:
        If left as False, `df.to_excel()` is used to write out df. However,
        this can cause compatibility issues with data validation in the Excel
        workbook. If this is the cases, set to True. If True, uses openpyxl
        functionality to write out `df` instead, reducing functionality of the
        write, but keeping the data validation.

    to_excel_kwargs:
        Any further kwargs to pass to `df.to_excel()`
        Ignored if `keep_data_validation` is True.

    Returns
    -------
    None
    """
    if keep_data_validation:
        return _openpyxl_df_to_excel(
            df=df,
            path=path,
            sheet_name=sheet_name,
            header=header,
            index=index,
            start_row=start_row,
            start_col=start_col,
        )

    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(path):
        df.to_excel(
            path,
            header=header,
            index=index,
            startrow=start_row,
            startcol=start_col,
            **to_excel_kwargs,
        )
        return

    # ignore engine parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    with pd.ExcelWriter(    # type: ignore[abstract]
        path,
        engine='openpyxl',
        mode='a',
        if_sheet_exists='overlay',
    ) as writer:
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            header=header,
            index=index,
            startrow=start_row,
            startcol=start_col,
            **to_excel_kwargs,
        )


def prepend_cols(
    df: pd.DataFrame,
    col_names: Iterable[Any],
    col_vals: Iterable[Any],
    allow_duplicates: bool = False,
) -> pd.DataFrame:
    """Prepend the given columns to the dataframe

    Adds the given columns to the start of the dataframe, in the order given.
    That is, if a dataframe has columns ["d", "e", "f"], and the `col_names`
    given are ["a", "b", "c"], the resultant dataframe would have columns in
    the following order:["a", "b", "c", "d", "e", "f"].

    Parameters
    ----------
    df:
        The original dataframe to prepend `col_names` to with `col_values`.

    col_names:
        The names to give to the columns being prepended to `df`.

    col_vals:
        The values to give to each `col_names` being added to `df`. The index
        if each value should match that in `col_names`. That is:
        `df[col_name[idx]] = col_vals[idx]`

    allow_duplicates:
        Whether to allow this function to add a column to df if one with the
        same name already exists.

    Returns
    -------
    prepended_df:
        The original `df` with the given columns prepended with their values.
    """
    df = df.copy()

    # Validate inputs
    if len(col_names) != len(col_vals):
        raise ValueError(
            "col_names and col_vals need to be the same length. Got lengths:\n"
            f"col_names: {len(col_names)}\n"
            f"col_vals: {len(col_vals)}"
        )

    # Attach all columns in the given order
    for name, val in zip(col_names, col_vals):
        df.insert(loc=0, column=name, value=val, allow_duplicates=allow_duplicates)

    return df
