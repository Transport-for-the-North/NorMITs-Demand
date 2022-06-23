from openpyxl import load_workbook
from counts import Constants
import pandas as pd
"""
Process to apply counts factors to counts
"""


#df = pd.DataFrame(index=constants.growth_cols)
wb = load_workbook(r"C:\Projects\MidMITS\counts\growth_stats_11.xlsx",read_only=True)
growth = pd.read_csv(r"C:\Projects\MidMITS\counts\stacked_district11NOM1.csv").groupby(['NAME','Time Period', 'Vehicle Type']).sum()
growth['Growth'] = (growth['21'] / growth['18'])
dash = pd.read_csv(r"C:\Projects\MidMITS\counts\county_output_18.csv")


for county in dash['County'].dropna().unique():
    #if district in wb.sheetnames:
    #sheet = pd.read_excel(r"C:\Projects\MidMITS\growth_stats_11.xlsx",sheet_name=district).set_index('Unnamed: 0')
    #df[district] = pd.Series(sheet.loc['mean'])
    cols = Constants.dash_cols
    for time in ['AM','IP','PM']:
        for type in ['Car', 'LGV', 'HGV']:
            #dash[dash['LAD21NM']==district][f"{column}_2021"] = dash[dash['LAD21NM']==district][f"{column}"] * (1+df.loc[cols[column],district])
            dash.loc[dash['County']==county,f"{time} {type} 2021"] = dash.loc[dash['County']==county,f"{time} {type}"] * (growth.loc[(county,time,type),'Growth'])
            print(f"{county}, {time}, {type}, {growth.loc[(county,time,type),'Growth']}")
    #else:
         #print(f"no data for {district}")
dash.set_index(['A','B']).to_csv(r"C:\Projects\MidMITS\counts\county_output_21.csv")
#df.to_csv(r"C:\Projects\MidMITS\counts\factors.csv")